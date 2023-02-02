import urllib.request
from shutil import copyfile
import openpyxl
import re
import os
from config import EXCEL_PATH
from ..data.sim import GBFSim
from module_huiji.danteng_downloader import Downloader


def find_index(colTitle, workSheet):
    for cell in workSheet[1]:
        if cell.value == colTitle:
            return cell.column_letter
    return 1


def merge_sheet(workSheet, excelListPath, ID_set, chsTitle, jpTitle):
    workBook2 = openpyxl.load_workbook(excelListPath)
    workSheet2 = workBook2.worksheets[0]
    ID_index2 = find_index('ID', workSheet2)
    Title_index2 = find_index(chsTitle, workSheet2)
    Name_index2 = find_index(jpTitle, workSheet2)
    for cell in workSheet2[ID_index2][2:]:
        characterID = cell.value
        if not isinstance(characterID, int) or characterID <= 0 or characterID in ID_set:
            continue
        row = cell.row-1
        newRow = [characterID]
        if workSheet2[Title_index2][row].value:
            newRow.append(workSheet2[Title_index2][row].value)
        elif workSheet2[Name_index2][row].value:
            newRow.append(workSheet2[Name_index2][row].value)
        else:
            newRow.append('')

        workSheet.append(newRow)

# 检查终突


def check_evo(workSheet, excelListPath, ID_set):
    workBook2 = openpyxl.load_workbook(excelListPath)
    workSheet2 = workBook2.worksheets[0]
    evo_index2 = find_index('max_evo', workSheet2)
    ID_index2 = find_index('ID', workSheet2)
    extra_index = find_index('extra_ver', workSheet)
    for cell in workSheet2[evo_index2][2:]:
        max_evo = cell.value
        if not isinstance(max_evo, int) or max_evo < 5:
            continue
        row = cell.row-1
        characterID = workSheet2[ID_index2][row].value or 0
        if ID_set[characterID][1]:
            continue
        workSheet[extra_index][ID_set[characterID][0]].value = '03'


# 检测角色列表和皮肤excel文件，将新角色整合进角色动画excel文件中
def merge_anime_excel(cfg, args):
    excelAnimePath = EXCEL_PATH+'/角色动画.xlsx'
    excelListPath = EXCEL_PATH+'/角色列表.xlsx'
    excelSkinPath = EXCEL_PATH+'/皮肤.xlsx'

    workBook = openpyxl.load_workbook(excelAnimePath)
    workSheet = workBook.worksheets[0]
    ID_index = find_index('ID', workSheet)
    extra_index = find_index('extra_ver', workSheet)

    ID_set = {}
    for row in range(1, workSheet.max_row):
        ID_set[workSheet[ID_index][row].value] = (
            row, workSheet[extra_index][row].value)

    merge_sheet(workSheet, excelListPath, ID_set, 'tag_title', 'name_chs')
    merge_sheet(workSheet, excelSkinPath, ID_set, 'name_chs', 'name_jp')
    check_evo(workSheet, excelListPath, ID_set)
    try:
        workBook.save(excelAnimePath)
    except:
        print('动画Excel文件保存失败')
        return False


def update_anime_page(cfg, args):
    jsUploadPath = 'anime/upload/'
    gbf_wiki = cfg['wiki']
    pathList = ['model/manifest/', 'cjs/', ]

    for filesPath in pathList:
        if not os.path.exists(jsUploadPath+filesPath):
            continue
        for fileName in os.listdir(jsUploadPath+filesPath):
            if len(fileName) <= 3 or fileName[-3:] != '.js':
                continue
            page_title = f'Js:{filesPath}{fileName}'
            try:
                openedFile = open(
                    jsUploadPath+filesPath+fileName, "r")
                page_content = openedFile.read()
                openedFile.close()
            except:
                if openedFile:
                    openedFile.close()
                continue

            gbf_wiki.edit(page_title, page_content)
            # 上传后删除文件
            os.remove(jsUploadPath+filesPath+fileName)

        gbf_wiki.wait_threads()


def check_anime_file(cfg, args):
    gbf_sim = GBFSim(cfg)
    gameVersion = gbf_sim._version
    excelFile = EXCEL_PATH+'/角色动画.xlsx'
    excelFile2 = EXCEL_PATH+'/角色动画变身.xlsx'
    imgUploadPath = 'UPLOAD/'
    rootPath = 'anime/'
    jsUploadPath = 'anime/upload/'
    pathManifest = 'anime/model/manifest/'
    pathCjs = 'anime/cjs/'
    pathImg = 'anime/image/'
    mode = 0  # 0的保存格式为wiki使用，1的保存格式为本地使用
    if not os.path.exists(pathManifest):
        os.makedirs(pathManifest)
    if not os.path.exists(pathCjs):
        os.makedirs(pathCjs)
    if not os.path.exists(pathImg):
        os.makedirs(pathImg)
    if not os.path.exists(jsUploadPath):
        os.makedirs(jsUploadPath)

    workBook = openpyxl.load_workbook(excelFile)
    workSheet = workBook.worksheets[0]
    maxRow = workSheet.max_row
    maxCol = workSheet.max_column
    toIndex = {}  # 列名称转实际目录
    typeList = []
    nspSuffixList = ['b', 'c', 'd', 'e', 'f', 'g', 'h',
                     'i', 'j', 'k', ]  # 奥义动画可能的后缀，可以参考R角色多萝赛尔
    pattern = '/cjs.*?\.png'

    '''
    变身版所需
    indexTransformMap = {}
    f1 = '_f1'
    '''

    # 开关
    retry_times = 5
    if 'IMAGE' in cfg:
        if 'retry' in cfg['IMAGE']:
            try:
                retry_times = int(cfg['IMAGE']['retry'])
            except:
                pass

    # 配置下载器
    downloader = Downloader()
    downloader.set_try_count(retry_times)

    for col in range(maxCol):           # 初始化toIndex
        toIndex[workSheet[1]
                [col].value] = openpyxl.utils.get_column_letter(col+1)

    IDIndex = toIndex['ID']
    npcIndex = toIndex['npc_01']
    extraIndex = toIndex['extra_ver']
    nameIndex = toIndex['name_chs']
    unusedIndex = toIndex['unused']
    tupleNpcIndex = tuple(workSheet[npcIndex])
    tupleExtraIndex = tuple(workSheet[extraIndex])

    retry_times = 5
    if 'IMAGE' in cfg:
        if 'retry' in cfg['IMAGE']:
            try:
                retry_times = int(cfg['IMAGE']['retry'])
            except:
                pass

    def save_file(url, fileName):
        f = open(fileName, "wb")
        content = url.read()
        f.write(content)
        f.close()

    # ——————————需要增加多线程下载———————————
    def download_js(head, charID, row, col):
        print('下载'+head+charID+'：', end='')
        # 下载动画文件cjs
        link = 'https://prd-game-a3-granbluefantasy.akamaized.net/assets/' + \
            gameVersion+'/js/cjs/'+head+charID+'.js'
        try:
            url = urllib.request.urlopen(link)
            fileName = 'cjs/'+head+charID+'.js'
            if not os.path.exists(jsUploadPath+'cjs'):
                os.makedirs(jsUploadPath+'cjs')
            save_file(url, jsUploadPath+fileName)
            copyfile(jsUploadPath+fileName, rootPath+fileName)
        except:
            print('文件不存在')
            return False

        # 下载配置文件manifest
        link = 'https://prd-game-a3-granbluefantasy.akamaized.net/assets/' + \
            gameVersion+'/js/model/manifest/'+head+charID+'.js'
        fileName = 'model/manifest/'+head+charID+'.js'
        downloader.download_multi_copies(
            link, [jsUploadPath+fileName, rootPath+fileName])
        '''
        try:
            url = urllib.request.urlopen(link)
            fileName = 'model/manifest/'+head+charID+'.js'
            if not os.path.exists(jsUploadPath+'model/manifest'):
                os.makedirs(jsUploadPath+'model/manifest')
            save_file(url, jsUploadPath+fileName)
            copyfile(jsUploadPath+fileName, rootPath+fileName)
        except:
            print('文件不存在')
            return False
            '''

        # 若下载成功则写入Excel，同一个位置可能存在多个文件，多余的部分记录下来之后人工判断实际使用的是哪个
        if not workSheet[col][row].value:
            workSheet[col][row].value = charID
        else:
            if workSheet[unusedIndex][row].value:
                unusedList = workSheet[unusedIndex][row].value.split(';')
                if head+charID in unusedList:
                    workSheet[unusedIndex][row].value = workSheet[unusedIndex][row].value+';'+head+charID
            else:
                workSheet[unusedIndex][row].value = head+charID
        print('成功')
        return True

    def download_img(charID):
        if mode == 0:
            fileName = 'Img-sp-cjs-'+charID+'.png'
        else:
            fileName = charID+'.png'
        link = 'https://prd-game-a1-granbluefantasy.akamaized.net/assets/img/sp/cjs/'+charID+'.png'
        downloader.download_multi_copies(
            link, [pathImg+fileName, imgUploadPath+fileName])
        return True

    for row in range(2, maxRow):
        # 检查extraIndex，跳过已填写的npcIndex
        if tupleNpcIndex[row].value:
            if tupleExtraIndex[row].value:
                typeList = tupleExtraIndex[row].value.split(';')
                for animeVersion in typeList:
                    if not workSheet[toIndex['npc_'+animeVersion]][row].value:
                        break
                else:
                    typeList = []
                    continue
            else:
                continue

        charID = str(workSheet[IDIndex][row].value)
        if charID == '0':
            continue

        # 生成待检查列表
        if workSheet[nameIndex][row].value:
            print('检查'+charID+' - '+workSheet[nameIndex][row].value+' 文件')
        else:
            print('检查'+charID+'文件')
        typeList.append('01')
        if charID[0:3] != '371':    # 皮肤通常只有01号动画
            typeList.append('02')

        # 检查并下载文件
        for animeVersion in typeList:
            col = toIndex['npc_'+animeVersion]           # 检查npc文件（小人）
            if not workSheet[col][row].value:
                download_js('npc_', charID+'_'+animeVersion, row, col)
                download_js('npc_', charID+'_'+animeVersion+'_s2', row, col)

            col = toIndex['nsp_'+animeVersion]            # 检查nsp文件（奥义特效）
            if not workSheet[col][row].value:
                # s2为特效在屏幕固定位置，s3为特效在屏幕固定位置且图层在角色下面
                for tempName in ['', '_s2', '_s3']:
                    tempName = charID+'_'+animeVersion+tempName
                    if download_js('nsp_', tempName, row, col):
                        for alpha in nspSuffixList:                    # 检查可能的额外nsp文件
                            if not download_js('nsp_', tempName+'_'+alpha, row, col):
                                break

            col = toIndex['hit_'+animeVersion]           # 检查phit文件（平A特效）
            if not workSheet[col][row].value:
                if animeVersion == '01':
                    download_js('phit_', charID, row, col)
                download_js('phit_', charID+'_'+animeVersion, row, col)
        typeList = []
        try:
            workBook.save(excelFile)
        except:
            pass
        downloader.wait_threads()

    # 保存Excel
    try:
        workBook.save(excelFile)
    except:
        print('动画Excel文件保存失败')
        return False

    # 下载图片
    for fileName in os.listdir(jsUploadPath+'model/manifest/'):
        if len(fileName) <= 3 or fileName[-3:] != '.js':
            continue
        manifestFile = open(jsUploadPath+'model/manifest/'+fileName, "r")
        manifestText = manifestFile.read()
        imgPatternList = [(m.start()+5, m.end()-4)
                          for m in re.finditer(pattern, manifestText)]
        imgNameList = [0]*len(imgPatternList)
        for i, imgSet in enumerate(imgPatternList):
            imgNameList[i] = manifestText[imgSet[0]:imgSet[1]]
        manifestFile.close()
        for imgName in imgNameList:
            download_img(imgName)
    downloader.wait_threads()

    return True
